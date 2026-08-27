"""
bom_reader.py
-------------
Reads an Excel (.xlsx) file:
  1. Scans ROW 1 left-to-right to locate Description, Part Type, Qty, Unit columns
     using the alias list in component_dictionary.json.
  2. Scans all data rows (row 2 onward) and returns structured dicts.
"""

import json
import os
import re
import openpyxl

# --- HOTFIX FOR ERP EXPORTS ---
# Many ERP systems (like HPCC DB) generate malformed styles.xml that crash openpyxl.
# Since we only care about data, we bypass the entire stylesheet parsing.
import openpyxl.reader.excel
openpyxl.reader.excel.apply_stylesheet = lambda archive, wb: None
# ------------------------------

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))

from file_versions import active_filename as _active_filename

with open(os.path.join(_BASE_DIR, _active_filename("component_dictionary.json")),
          encoding="utf-8") as f:
    _DICT = json.load(f)

_ALIASES = _DICT["column_header_aliases"]


def _norm_header(value) -> str:
    """Normalize a header cell: lowercase, strip all non-alphanumeric chars.
    'MFR. P/N' -> 'mfrpn', 'Qty.' -> 'qty'."""
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


# Build reverse lookup: normalized header text → role name
_HEADER_ROLE: dict[str, str] = {}
for role, aliases in _ALIASES.items():
    if role.startswith("_"):
        continue
    for alias in aliases:
        _HEADER_ROLE.setdefault(_norm_header(alias), role)

# Aliases that exactly name their role take priority over generic ones
_PRIMARY_ALIASES: dict[str, set] = {
    "partnumber_col": {"partnumber", "partnumbercol"},
}


def detect_columns(headers: list) -> dict[str, int]:
    """
    Scan header list (row 1) left-to-right.
    Returns dict: role → column index (0-based).
    E.g. {"description_col": 3, "part_type_col": 2, ...}
    A later column with a higher-priority alias overrides an earlier one.
    """
    found: dict[str, int] = {}
    prio: dict[str, int] = {}
    for col_idx, cell_value in enumerate(headers):
        if cell_value is None:
            continue
        normalized = _norm_header(cell_value)
        if not normalized:
            continue
        role = _HEADER_ROLE.get(normalized)
        if role is None:
            continue
        # Primary aliases (exact role name) beat generic ones like 'item number'
        p = 2 if normalized in _PRIMARY_ALIASES.get(role, set()) else 1
        if role not in found or p > prio[role]:   # higher priority wins
            found[role] = col_idx
            prio[role] = p
    return found


def read_bom_file(filepath: str) -> tuple[list[dict], dict, str | None, list]:
    """
    Open xlsx file. For each sheet, try to detect BOM columns in row 1.
    Returns:
      (rows, col_map, sheet_name, header_values)
        rows        : list of dicts per data row
        col_map     : {role: col_index}
        sheet_name  : name of the sheet used
        header_values: list of header cell values from the detected header row
    Returns ([], {}, None, []) if no suitable sheet found.

    Each row dict has keys:
      raw_row      : tuple of all cell values
      description  : str or None
      part_type    : str or None
      quantity     : float or None
      unit         : str or None
      partnumber   : str or None
      row_number   : int (1-based excel row)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    best_sheet = None
    best_col_map: dict[str, int] = {}
    best_score = 0
    best_data_start_row = 2

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is not None and ws.max_row < 2:
            continue
        
        # Scan the first 30 rows to dynamically find the header row
        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=30), start=1):
            row_values = [cell.value for cell in row_cells]
            col_map = detect_columns(row_values)
            score = len(col_map)
            
            # Bonus: must have description col to be useful
            if "description_col" in col_map:
                score += 10
                
            if score > best_score:
                best_score = score
                best_sheet = sheet_name
                best_col_map = col_map
                best_data_start_row = row_idx + 1

    if not best_sheet or "description_col" not in best_col_map:
        return [], {}, None, []

    # Get header row values from the detected header row
    header_row_idx = best_data_start_row - 1
    header_ws = wb[best_sheet]
    header_values = []
    for cell in header_ws[header_row_idx]:
        header_values.append(cell.value)

    ws = wb[best_sheet]
    rows_out: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=best_data_start_row, values_only=True), start=best_data_start_row):
        def _get(role):
            idx = best_col_map.get(role)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return val

        desc = _get("description_col")
        if desc is None:
            continue                   # skip blank description rows
        desc_str = str(desc).strip()
        if not desc_str:
            continue

        pt_raw = _get("part_type_col")
        qty_raw = _get("quantity_col")
        unit_raw = _get("unit_col")
        pn_raw = _get("partnumber_col")
        mpn_raw = _get("mpn_col")
        ipn_raw = _get("internal_pn_col")

        # Parse qty safely
        qty = None
        if qty_raw is not None:
            try:
                qty = float(qty_raw)
            except (ValueError, TypeError):
                qty = None

        def _clean(val):
            return str(val).strip() if val is not None and str(val).strip() else None

        pn_generic = _clean(pn_raw)
        rows_out.append({
            "raw_row":     row,
            "description": desc_str,
            "part_type":   str(pt_raw).strip() if pt_raw else None,
            "quantity":    qty,
            "unit":        str(unit_raw).strip() if unit_raw else None,
            # Generic 'Part Number' is mostly the MFR code; also used as
            # Internal P/N when no dedicated column exists
            "partnumber":  pn_generic,
            "mpn":         _clean(mpn_raw) or pn_generic,
            "internal_pn": _clean(ipn_raw) or pn_generic,
            "row_number":  row_idx,
        })

    return rows_out, best_col_map, best_sheet, header_values


def get_available_sheets(filepath: str) -> list[str]:
    """Return list of sheet names in the workbook."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names
